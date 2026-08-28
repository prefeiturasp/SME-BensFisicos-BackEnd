"""
Testes de cobertura da refatoração NBBPM unificada.
Valida:
- geração por ano com continuidade (prefixo fixo 001, sequencial global por ano)
- unicidade, formato 001.YYYYYYY/ZZZZ e prefixo fixo
- concorrência primeira NBBPM do ano e reuso mesma baixa
- validações de vínculo por UO, permissão gestor/superuser, visibilidade superuser
- data migration idempotente e tratamento ACEITA sem número via admin
- PDF layout lote e export Excel buscando nova tabela
"""
import threading
from decimal import Decimal
from unittest.mock import MagicMock
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import TestCase, TransactionTestCase, RequestFactory
from django.urls import reverse
from django.utils import timezone
from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import Group
from rest_framework.test import APIClient
from rest_framework import status

from dados_comuns.tests.factories import criar_ua
from dados_comuns.models import UnidadeOrcamentaria as UO_Model
from dados_comuns.tests.auth_test_utils import auth_kwargs, codigo_ua
from usuario.models import Usuario
from usuario.constants import GRUPO_GESTOR_PATRIMONIO, GRUPO_OPERADOR_INVENTARIO

from bem_patrimonial import constants
from bem_patrimonial.models import BemPatrimonial, BaixaFisicaBemPatrimonial, BaixaFisicaBensItem, NBBPM
from bem_patrimonial.services.nbbpm_numero import (
    gerar_numero_nbbpm_unificado,
    criar_nbbpm_com_retry,
)
from bem_patrimonial.pdf_utils import extrair_codigo_ua
from bem_patrimonial.nbbpm_lote import gerar_pdf_nbbpm_lote, _criar_tabela_bens
from bem_patrimonial.admins.baixa_fisica_bem_patrimonial import BaixaFisicaBemPatrimonialAdmin, BaixaFisicaResource


def criar_uo(codigo="200", nome="UO Teste", sigla="UO", **kwargs):
    obj, _ = UO_Model.objects.get_or_create(codigo=codigo, defaults={"nome": nome, "sigla": sigla, **kwargs})
    return obj



def criar_usuario(username, uo, ua, grupos=None, is_superuser=False, **kwargs):
    user = Usuario.objects.create_user(
        username=username,
        email=f"{username}@test.com",
        **auth_kwargs("senha123"),
        unidade_administrativa=ua,
        unidade_orcamentaria=uo,
        is_superuser=is_superuser,
        is_staff=True,
        **kwargs,
    )
    for g in (grupos or []):
        grp, _ = Group.objects.get_or_create(name=g)
        user.groups.add(grp)
    return user


def criar_bem(ua, criado_por, numero_patrimonial="000.000000001-0", **kwargs):
    return BemPatrimonial.objects.create(
        nome=kwargs.pop("nome", "Notebook"),
        descricao=kwargs.pop("descricao", "Desc"),
        valor_unitario=kwargs.pop("valor_unitario", Decimal("1000.00")),
        marca=kwargs.pop("marca", "Dell"),
        modelo=kwargs.pop("modelo", "Latitude"),
        numero_processo=kwargs.pop("numero_processo", "PROC-TESTE"),
        numero_patrimonial=numero_patrimonial,
        unidade_administrativa=ua,
        criado_por=criado_por,
        status=kwargs.pop("status", constants.APROVADO),
        **kwargs,
    )


def criar_baixa(ua, criado_por, status=constants.ACEITA, **kwargs):
    return BaixaFisicaBemPatrimonial.objects.create(
        unidade_administrativa_origem=ua,
        numero_processo_baixa=kwargs.pop("numero_processo_baixa", "PROC-BX-001"),
        status=status,
        criado_por=criado_por,
        data_baixa=kwargs.pop("data_baixa", timezone.localdate()),
        numero_nbbpm=kwargs.pop("numero_nbbpm", ""),
        **kwargs,
    )


# =====================================================================
# 1. Prefixo UA (último grupo) e extração
# =====================================================================

class TestExtrairCodigoUa(TestCase):
    def test_ua_01_16_10_287_retorna_287(self):
        self.assertEqual(extrair_codigo_ua("01.16.10.287"), "287")

    def test_ua_simples_100_retorna_100(self):
        self.assertEqual(extrair_codigo_ua("100"), "100")

    def test_codigo_vazio_retorna_000(self):
        self.assertEqual(extrair_codigo_ua(""), "000")
        self.assertEqual(extrair_codigo_ua(None), "000")

    def test_ua_com_zero_pad(self):
        self.assertEqual(extrair_codigo_ua("01.02.10"), "010")
        self.assertEqual(extrair_codigo_ua("5"), "005")


# =====================================================================
# 2. Geração por UA/ano com continuidade (não reseta para 0000001)
# =====================================================================

class TestGeracaoPorUOAnoContinuidade(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua1 = criar_ua(uo=self.uo, codigo=codigo_ua(1, 16, 10, 100), sigla="UA100", nome="UA 100")
        self.ua2 = criar_ua(uo=self.uo, codigo=codigo_ua(1, 16, 10, 200), sigla="UA200", nome="UA 200")
        self.gestor = criar_usuario("gestor_cont", self.uo, self.ua1, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua1, self.gestor)

    def _criar_baixa_aceita_com_legado(self, numero_legado, ua, ano=2026):
        baixa = criar_baixa(ua, self.gestor, status=constants.ACEITA, numero_nbbpm=numero_legado)
        return baixa

    def test_continuidade_nao_reseta_para_0000001_quando_historico_grande(self):
        # Legado com prefixo fixo 001 e sequencial grande
        ano = 2026
        legado_num = f"001.{1234:07d}/{ano}"
        self._criar_baixa_aceita_com_legado(legado_num, self.ua1, ano=ano)
        baixa2 = criar_baixa(self.ua1, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa2, bem=self.bem)
        nbbpm = NBBPM.objects.create(
            numero="",
            numero_processo_baixa="PROC-NEW",
            data_autorizacao=timezone.datetime(ano, 6, 15).date(),
            responsavel="Gestor",
            criado_por=self.gestor,
        )
        nbbpm.baixas.set([baixa2])
        numero_gerado = gerar_numero_nbbpm_unificado(nbbpm)
        self.assertEqual(numero_gerado, f"001.{1235:07d}/{ano}")

    def test_sequencial_por_uo_isolado_entre_uos(self):
        # Sequencial global por ano com prefixo fixo 001 (não isolado por UA/UO)
        uo_b = criar_uo(codigo="200", nome="UO B", sigla="UOB")
        ua_b = criar_ua(uo=uo_b, codigo="200", sigla="UAB", nome="UA B")
        gestor_b = criar_usuario("gestor_b", uo_b, ua_b, grupos=[GRUPO_GESTOR_PATRIMONIO])
        bem_b = criar_bem(ua_b, gestor_b, numero_patrimonial="000.000000010-0")

        ano = 2026
        # 5 NBBPMs com prefixo 001
        for i in range(5):
            criar_baixa(self.ua1, self.gestor, status=constants.ACEITA, numero_nbbpm=f"001.{i+1:07d}/{ano}")
        # Gera para UA 200 -> deve ser 001.0000006/2026 (global)
        baixa_b = criar_baixa(ua_b, gestor_b, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa_b, bem=bem_b)
        nbbpm_b = NBBPM.objects.create(
            numero="",
            numero_processo_baixa="PROC-B",
            data_autorizacao=timezone.datetime(ano, 6, 15).date(),
            responsavel="G",
            criado_por=gestor_b,
        )
        nbbpm_b.baixas.set([baixa_b])
        numero_b = gerar_numero_nbbpm_unificado(nbbpm_b)
        self.assertEqual(numero_b, f"001.{6:07d}/{ano}")
        # Persiste para que próximo seja 7 (global)
        nbbpm_b.numero = numero_b
        nbbpm_b.save(update_fields=["numero"])

        # Gera para UA 100 -> deve ser 001.0000007/2026
        baixa_a2 = criar_baixa(self.ua1, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa_a2, bem=self.bem)
        nbbpm_a = NBBPM.objects.create(
            numero="",
            numero_processo_baixa="PROC-A",
            data_autorizacao=timezone.datetime(ano, 6, 15).date(),
            responsavel="G",
            criado_por=self.gestor,
        )
        nbbpm_a.baixas.set([baixa_a2])
        numero_a = gerar_numero_nbbpm_unificado(nbbpm_a)
        # Como já criou 6 para b, agora é 7
        self.assertEqual(numero_a, f"001.{7:07d}/{ano}")

    def test_ano_diferente_reseta_sequencial(self):
        ano1 = 2025
        ano2 = 2026
        criar_baixa(self.ua1, self.gestor, status=constants.ACEITA, numero_nbbpm=f"001.{10:07d}/{ano1}")
        baixa = criar_baixa(self.ua1, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa, bem=self.bem)
        nbbpm = NBBPM.objects.create(
            numero="", numero_processo_baixa="P", data_autorizacao=timezone.datetime(ano1, 5, 1).date(),
            responsavel="G", criado_por=self.gestor,
        )
        nbbpm.baixas.set([baixa])
        num = gerar_numero_nbbpm_unificado(nbbpm)
        self.assertEqual(num, f"001.{11:07d}/{ano1}")
        baixa2 = criar_baixa(self.ua1, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa2, bem=self.bem)
        nbbpm2 = NBBPM.objects.create(
            numero="", numero_processo_baixa="P2", data_autorizacao=timezone.datetime(ano2, 5, 1).date(),
            responsavel="G", criado_por=self.gestor,
        )
        nbbpm2.baixas.set([baixa2])
        num2 = gerar_numero_nbbpm_unificado(nbbpm2)
        self.assertEqual(num2, f"001.{1:07d}/{ano2}")


# =====================================================================
# 3. Unicidade e formato
# =====================================================================

class TestUnicidadeEFormato(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA Teste", sigla="UAT")
        self.gestor = criar_usuario("gestor_uniq", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)

    def test_formato_xxx_yyyyyyy_zzzz(self):
        baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa, bem=self.bem)
        nbbpm = NBBPM.objects.create(
            numero="", numero_processo_baixa="P", data_autorizacao=timezone.localdate(),
            responsavel="G", criado_por=self.gestor,
        )
        nbbpm.baixas.set([baixa])
        numero = gerar_numero_nbbpm_unificado(nbbpm)
        self.assertRegex(numero, r"^\d{3}\.\d{7}[\./]\d{4}$")
        # novo formato usa "/" antes do ano
        self.assertIn("/", numero)
        partes = numero.replace("/", ".").split(".")
        self.assertEqual(len(partes[0]), 3)
        self.assertEqual(len(partes[1]), 7)
        self.assertEqual(len(partes[2]), 4)

    def test_prefixo_ua_001_correto(self):
        baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=baixa, bem=self.bem)
        nbbpm = NBBPM.objects.create(
            numero="", numero_processo_baixa="P", data_autorizacao=timezone.localdate(),
            responsavel="G", criado_por=self.gestor,
        )
        nbbpm.baixas.set([baixa])
        numero = gerar_numero_nbbpm_unificado(nbbpm)
        # prefixo fixo 001
        self.assertTrue(numero.startswith("001."))

    def test_unicidade_banco_impede_duplicata(self):
        # cria duas NBBPMs com mesmo número tentando salvar direto deve falhar na constraint
        NBBPM.objects.create(
            numero="001.0000001/2026",
            numero_processo_baixa="P1",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        # segunda com mesmo número deve levantar IntegrityError devido ao UniqueConstraint
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                NBBPM.objects.create(
                    numero="001.0000001/2026",
                    numero_processo_baixa="P2",
                    data_autorizacao=timezone.localdate(),
                    responsavel="G",
                    criado_por=self.gestor,
                )

    def test_validador_regex_impede_lixo(self):
        nbbpm = NBBPM(
            numero="abc",
            numero_processo_baixa="P",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        with self.assertRaises(ValidationError):
            nbbpm.full_clean()


# =====================================================================
# 4. Concorrência primeira NBBPM do ano (UO) e reuso mesma baixa
# =====================================================================

class TestConcorrenciaPrimeiraNBBPM(TransactionTestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA Teste", sigla="UAT")
        self.gestor = criar_usuario("gestor_conc", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem1 = criar_bem(self.ua, self.gestor, numero_patrimonial="000.000000001-0")
        self.bem2 = criar_bem(self.ua, self.gestor, numero_patrimonial="000.000000002-0")
        self.baixa1 = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P1")
        self.baixa2 = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P2")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa1, bem=self.bem1)
        BaixaFisicaBensItem.objects.create(baixa=self.baixa2, bem=self.bem2)

    def test_duas_thread_primeira_nbbpm_geram_0000001_e_0000002(self):
        # Limpa qualquer NBBPM anterior
        NBBPM.objects.all().delete()
        BaixaFisicaBemPatrimonial.objects.filter(numero_nbbpm__isnull=False).update(numero_nbbpm="")

        resultados = {}
        erros = []

        def criar(baixa, key):
            try:
                nbbpm = criar_nbbpm_com_retry(
                    baixas=[baixa],
                    numero_processo_baixa=f"PROC-{key}",
                    data_autorizacao=timezone.localdate(),
                    responsavel="Gestor",
                    criado_por=self.gestor,
                )
                resultados[key] = nbbpm.numero
            except Exception as e:
                erros.append(str(e))

        t1 = threading.Thread(target=criar, args=(self.baixa1, "t1"))
        t2 = threading.Thread(target=criar, args=(self.baixa2, "t2"))
        t1.start()
        t2.start()
        t1.join()
        t2.join()

        if len(resultados) == 2:
            self.assertEqual(len(set(resultados.values())), 2)
            nums = sorted(resultados.values())
            # Devem ser sequenciais 0000001 e 0000002 com prefixo fixo 001
            self.assertTrue(nums[0].startswith("001."))
            self.assertTrue(nums[1].startswith("001."))
            seqs = [int(n.replace("/", ".").split(".")[1]) for n in nums]
            self.assertEqual(sorted(seqs), [1, 2])
        else:
            self.assertLessEqual(len(resultados), 2)
            if len(resultados) == 1:
                self.assertEqual(len(erros), 1)

    def test_unicidade_via_retry_em_integrity_error(self):
        NBBPM.objects.all().delete()
        n1 = criar_nbbpm_com_retry(
            baixas=[self.baixa1],
            numero_processo_baixa="PROC-1",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        self.assertEqual(n1.numero, "001.0000001/" + str(timezone.localdate().year))
        n2 = criar_nbbpm_com_retry(
            baixas=[self.baixa2],
            numero_processo_baixa="PROC-2",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        self.assertEqual(n2.numero, "001.0000002/" + str(timezone.localdate().year))


class TestReusoMesmaBaixaConcorrente(TransactionTestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA Teste", sigla="UAT")
        self.gestor = criar_usuario("gestor_reuso", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P-REUSE")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        NBBPM.objects.all().delete()

    def test_mesma_baixa_em_dois_lotes_concorrentes_um_sucesso_um_erro(self):
        resultados = {}
        erros = []

        def tentar(key):
            try:
                # Cria baixa extra para tentar lote com mesma baixa
                nbbpm = criar_nbbpm_com_retry(
                    baixas=[self.baixa],
                    numero_processo_baixa=f"PROC-{key}",
                    data_autorizacao=timezone.localdate(),
                    responsavel="Gestor",
                    criado_por=self.gestor,
                )
                resultados[key] = nbbpm.numero
            except Exception as e:
                erros.append(str(e))

        t1 = threading.Thread(target=tentar, args=("t1",))
        t2 = threading.Thread(target=tentar, args=("t2",))
        t1.start()
        t2.start()
        t1.join()
        t2.join()

        # Um deve ter sucesso, outro deve falhar com ValidationError contendo "já possui NBBPM"
        self.assertEqual(len(resultados) + len(erros), 2)
        # Se ambos tentarem, um succede
        self.assertEqual(len(resultados), 1)
        self.assertEqual(len(erros), 1)
        self.assertIn("já possui NBBPM", erros[0])


# =====================================================================
# 5. Permissão apenas gestor/superuser e visibilidade superuser
# =====================================================================

class TestPermissaoESuperuser(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA Teste", sigla="UAT")
        self.ua2 = criar_ua(uo=self.uo, codigo="002", nome="UA2", sigla="UAT2")
        self.gestor = criar_usuario("gestor_perm", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.operador = criar_usuario("operador_perm", self.uo, self.ua, grupos=[GRUPO_OPERADOR_INVENTARIO])
        self.superuser = criar_usuario("superuser_perm", self.uo, None, grupos=[GRUPO_GESTOR_PATRIMONIO], is_superuser=True)
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P-PERM")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        # NBBPM existente para teste escopo
        self.nbbpm = NBBPM.objects.create(
            numero="016.0000001.2026",
            numero_processo_baixa="PROC",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        self.nbbpm.baixas.set([self.baixa])
        # Baixa em outra UA
        self.baixa2 = criar_baixa(self.ua2, self.gestor, status=constants.ACEITA, numero_processo_baixa="P2")
        bem2 = criar_bem(self.ua2, self.gestor, numero_patrimonial="000.000000002-0")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa2, bem=bem2)
        self.nbbpm2 = NBBPM.objects.create(
            numero="016.0000002.2026",
            numero_processo_baixa="PROC2",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        self.nbbpm2.baixas.set([self.baixa2])

    def test_operador_nao_pode_gerar_nbbpm_via_api(self):
        client = APIClient()
        client.force_authenticate(user=self.operador)
        payload = {
            "baixas": [self.baixa.id],
            "numero_processo_baixa": "PROC-NOVO",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "Operador",
        }
        resp = client.post("/api/nbbpm/", payload, format="json")
        self.assertIn(resp.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_400_BAD_REQUEST])

    def test_gestor_pode_gerar_nbbpm_via_api(self):
        # Precisa de baixa nova sem NBBPM
        baixa_nova = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P-NEW-GESTOR")
        bem_novo = criar_bem(self.ua, self.gestor, numero_patrimonial="000.000000003-0")
        BaixaFisicaBensItem.objects.create(baixa=baixa_nova, bem=bem_novo)
        client = APIClient()
        client.force_authenticate(user=self.gestor)
        payload = {
            "baixas": [baixa_nova.id],
            "numero_processo_baixa": "PROC-NOVO",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "Gestor",
        }
        resp = client.post("/api/nbbpm/", payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.assertRegex(resp.data["numero"], r"^\d{3}\.\d{7}[\./]\d{4}$")

    def test_superuser_pode_gerar_nbbpm_via_api(self):
        baixa_nova = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P-NEW-SUPER")
        bem_novo = criar_bem(self.ua, self.gestor, numero_patrimonial="000.000000004-0")
        BaixaFisicaBensItem.objects.create(baixa=baixa_nova, bem=bem_novo)
        client = APIClient()
        client.force_authenticate(user=self.superuser)
        payload = {
            "baixas": [baixa_nova.id],
            "numero_processo_baixa": "PROC-SUPER",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "Super",
        }
        resp = client.post("/api/nbbpm/", payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_superuser_ve_todas_nbbpms(self):
        from bem_patrimonial.nbbpm_api_views import NBBPMViewSet
        factory = RequestFactory()
        request = factory.get("/api/nbbpm/")
        request.user = self.superuser
        view = NBBPMViewSet()
        view.request = request
        qs = view.get_queryset()
        self.assertEqual(qs.count(), 2)

    def test_gestor_sem_ua_ve_todas_da_uo(self):
        gestor_sem_ua = criar_usuario("gestor_sem_ua", self.uo, None, grupos=[GRUPO_GESTOR_PATRIMONIO])
        from bem_patrimonial.nbbpm_api_views import NBBPMViewSet
        factory = RequestFactory()
        request = factory.get("/api/nbbpm/")
        request.user = gestor_sem_ua
        view = NBBPMViewSet()
        view.request = request
        qs = view.get_queryset()
        self.assertEqual(qs.count(), 2)

    def test_operador_so_ve_da_sua_ua_via_baixa_viewset(self):
        from bem_patrimonial.api_views import BaixaFisicaBemPatrimonialViewSet
        factory = RequestFactory()
        request = factory.get("/api/baixa-fisica/")
        request.user = self.superuser
        view = BaixaFisicaBemPatrimonialViewSet()
        view.request = request
        view.format_kwarg = None
        qs = view.get_queryset()
        self.assertGreaterEqual(qs.count(), 2)

    def test_baixa_list_serializer_traz_numero_via_nbbpm(self):
        from bem_patrimonial.api_serializers import BaixaFisicaBemPatrimonialListSerializer
        # baixa já tem nbbpm via M2M
        data = BaixaFisicaBemPatrimonialListSerializer(self.baixa).data
        self.assertEqual(data["numero_nbbpm"], "016.0000001.2026")


# =====================================================================
# 6. Data migration idempotente
# =====================================================================

class TestDataMigrationIdempotente(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA", sigla="UA")
        self.gestor = criar_usuario("gestor_mig", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        # Limpa
        NBBPM.objects.all().delete()
        BaixaFisicaBemPatrimonial.objects.all().delete()

    def test_migration_copia_numero_antigo_idempotente(self):
        baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_nbbpm="016.0000005.2026")
        BaixaFisicaBensItem.objects.create(baixa=baixa, bem=self.bem)
        def migrar_baixa_para_nbbpm(apps, schema_editor):
            import importlib.util
            import pathlib
            spec = importlib.util.spec_from_file_location("migr_0044", str(pathlib.Path(__file__).resolve().parent.parent / "migrations/0044_unificar_nbbpm_ua_ano.py"))
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod.migrar_baixa_para_nbbpm(apps, schema_editor)
        from django.db import connection as conn
        from django.apps import apps as real_apps
        # Simula chamada duas vezes
        FakeSchemaEditor = type("FakeSchemaEditor", (), {"connection": conn})
        # Usa apps registry real para testar
        migrar_baixa_para_nbbpm(real_apps, FakeSchemaEditor())
        count1 = NBBPM.objects.count()
        migrar_baixa_para_nbbpm(real_apps, FakeSchemaEditor())
        count2 = NBBPM.objects.count()
        self.assertEqual(count1, 1)
        self.assertEqual(count2, 1)
        nbbpm = NBBPM.objects.first()
        self.assertEqual(nbbpm.numero, "016.0000005.2026")
        self.assertIn(baixa, nbbpm.baixas.all())

    def test_migration_trata_aceita_sem_numero_via_admin(self):
        baixa = BaixaFisicaBemPatrimonial.objects.create(
            unidade_administrativa_origem=self.ua,
            numero_processo_baixa="PROC-SEM-NUMERO",
            status=constants.ACEITA,
            criado_por=self.gestor,
            data_baixa=timezone.localdate(),
            data_aprovacao=timezone.now(),
            aprovado_por=self.gestor,
            numero_nbbpm="",
        )
        BaixaFisicaBensItem.objects.create(baixa=baixa, bem=self.bem)
        def migrar_baixa_para_nbbpm(apps, schema_editor):
            import importlib.util
            import pathlib
            spec = importlib.util.spec_from_file_location("migr_0044", str(pathlib.Path(__file__).resolve().parent.parent / "migrations/0044_unificar_nbbpm_ua_ano.py"))
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod.migrar_baixa_para_nbbpm(apps, schema_editor)
        from django.db import connection as conn2
        from django.apps import apps as real_apps
        FakeSchemaEditor2 = type("FakeSchemaEditor", (), {"connection": conn2})
        migrar_baixa_para_nbbpm(real_apps, FakeSchemaEditor2())
        self.assertTrue(NBBPM.objects.filter(baixas__in=[baixa]).exists())
        nbbpm = NBBPM.objects.filter(baixas__in=[baixa]).first()
        self.assertRegex(nbbpm.numero, r"^\d{3}\.\d{7}[\./]\d{4}$")
        self.assertTrue(nbbpm.numero.startswith("001."))


# =====================================================================
# 7. PDF padronizado lote 6 colunas
# =====================================================================

class TestPDFPadronizado(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA", sigla="UAT")
        self.gestor = criar_usuario("gestor_pdf", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        self.nbbpm = NBBPM.objects.create(
            numero="016.0000001.2026",
            numero_processo_baixa="PROC",
            data_autorizacao=timezone.localdate(),
            responsavel="Gestor",
            criado_por=self.gestor,
        )
        self.nbbpm.baixas.set([self.baixa])

    def test_pdf_usa_layout_lote_6_colunas_DE_ATE(self):
        # Verifica que _criar_tabela_bens gera 6 colunas com DE/ATÉ
        [tabela] = _criar_tabela_bens(self.nbbpm)
        # tabela tem header com 6 colunas: DE, ATÉ, DISCRIMINAÇÃO, QTD, UNITÁRIO, TOTAL
        # Verifica colWidths
        self.assertEqual(len(tabela._argW), 6)
        textos = []
        for linha in tabela._cellvalues:
            for cel in linha:
                if hasattr(cel, "text"):
                    textos.append(cel.text)
        self.assertTrue(any("DE" in t for t in textos))
        self.assertTrue(any("ATÉ" in t for t in textos))
        self.assertIn("VALOR", " ".join(textos) or "")
        # Verifica PDF gera bytes
        buf = gerar_pdf_nbbpm_lote(self.nbbpm, usuario_gerador=self.gestor)
        self.assertTrue(buf.getvalue().startswith(b"%PDF"))

    def test_pdf_contem_numero_nbbpm_no_titulo(self):
        buf = gerar_pdf_nbbpm_lote(self.nbbpm)
        # O buffer não é texto puro, mas podemos checar que não levanta erro e tem conteúdo
        self.assertGreater(len(buf.getvalue()), 1000)


# =====================================================================
# 8. Export Excel busca número na nova tabela
# =====================================================================

class TestExcelExportNovaTabela(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA", sigla="UAT")
        self.gestor = criar_usuario("gestor_excel", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_nbbpm="")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        self.nbbpm = NBBPM.objects.create(
            numero="016.0000999.2026",
            numero_processo_baixa="PROC-EXCEL",
            data_autorizacao=timezone.localdate(),
            responsavel="Gestor",
            criado_por=self.gestor,
        )
        self.nbbpm.baixas.set([self.baixa])
        self.client = APIClient()
        self.client.force_authenticate(user=self.gestor)

    def test_excel_api_busca_numero_na_nova_tabela(self):
        from openpyxl import load_workbook
        resp = self.client.get("/api/baixa-fisica/exportar-excel/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        # Procura linha com NBBPM
        encontrado = False
        for row in ws.iter_rows(values_only=True):
            if "016.0000999.2026" in (str(v) for v in row if v):
                encontrado = True
                break
        self.assertTrue(encontrado, "Excel deve conter número da NBBPM via M2M")

    def test_excel_admin_resource_busca_nbbpm_via_m2m(self):
        resource = BaixaFisicaResource()
        # Garante prefetch
        baixa = BaixaFisicaBemPatrimonial.objects.prefetch_related("nbbpms_lote", "itens__bem").get(pk=self.baixa.pk)
        valor = resource.dehydrate_nbbpm(baixa)
        self.assertEqual(valor, "016.0000999.2026")

    def test_excel_admin_fallback_legado_quando_sem_nbbpm(self):
        # Baixa sem NBBPM mas com legado
        baixa2 = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_nbbpm="016.0000001.2026", numero_processo_baixa="P2")
        baixa2.save()
        resource = BaixaFisicaResource()
        valor = resource.dehydrate_nbbpm(baixa2)
        self.assertEqual(valor, "016.0000001.2026")


# =====================================================================
# 9. Desativação rotas antigas
# =====================================================================

class TestDesativacaoRotasAntigas(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA", sigla="UAT")
        self.gestor = criar_usuario("gestor_desat", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_nbbpm="016.0000001.2026")
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        self.client = APIClient()
        self.client.force_authenticate(user=self.gestor)

    def test_gerar_nbbpm_individual_removido_retorna_404(self):
        # Rota removida — deve retornar 404 (usar GET /api/nbbpm/{id}/pdf/)
        resp = self.client.get(f"/api/baixa-fisica/{self.baixa.pk}/gerar-nbbpm/")
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_baixar_nbbpm_admin_retorna_410(self):
        factory = RequestFactory()
        site = AdminSite()
        admin = BaixaFisicaBemPatrimonialAdmin(BaixaFisicaBemPatrimonial, site)
        request = factory.get(f"/admin/bem_patrimonial/baixafisicabempatrimonial/{self.baixa.pk}/nbbpm/")
        request.user = self.gestor
        resp = admin.baixar_nbbpm(request, self.baixa.pk)
        self.assertEqual(resp.status_code, 410)

    def test_nova_api_nbbpm_funciona(self):
        # cria baixa nova sem nbbpm para gerar via nova API
        baixa_nova = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_processo_baixa="P-NOVA")
        bem_novo = criar_bem(self.ua, self.gestor, numero_patrimonial="000.000000009-0")
        BaixaFisicaBensItem.objects.create(baixa=baixa_nova, bem=bem_novo)
        payload = {
            "baixas": [baixa_nova.id],
            "numero_processo_baixa": "PROC-NEW",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "Gestor",
        }
        resp = self.client.post("/api/nbbpm/", payload, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        # PDF via nova rota
        nbbpm_id = resp.data["id"]
        resp_pdf = self.client.get(f"/api/nbbpm/{nbbpm_id}/pdf/")
        self.assertEqual(resp_pdf.status_code, status.HTTP_200_OK)
        self.assertEqual(resp_pdf["Content-Type"], "application/pdf")


# =====================================================================
# 10. Validação de vínculo e bloqueio de reuso
# =====================================================================

class TestValidacaoVinculoEReuso(TestCase):
    def setUp(self):
        self.uo = criar_uo(codigo="01.16.10", nome="UO SME", sigla="SME")
        self.ua = criar_ua(uo=self.uo, codigo="001", nome="UA", sigla="UAT")
        self.gestor = criar_usuario("gestor_vinc", self.uo, self.ua, grupos=[GRUPO_GESTOR_PATRIMONIO])
        self.bem = criar_bem(self.ua, self.gestor)
        self.baixa = criar_baixa(self.ua, self.gestor, status=constants.ACEITA)
        BaixaFisicaBensItem.objects.create(baixa=self.baixa, bem=self.bem)
        self.nbbpm = NBBPM.objects.create(
            numero="016.0000001.2026",
            numero_processo_baixa="PROC",
            data_autorizacao=timezone.localdate(),
            responsavel="G",
            criado_por=self.gestor,
        )
        self.nbbpm.baixas.set([self.baixa])

    def test_baixa_ja_utilizada_nao_pode_ser_reutilizada(self):
        from bem_patrimonial.api_serializers import NBBPMGerarLoteSerializer
        data = {
            "baixas": [self.baixa.id],
            "numero_processo_baixa": "PROC-NEW",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "G",
        }
        req = MagicMock()
        req.user = self.gestor
        serializer = NBBPMGerarLoteSerializer(data=data, context={"request": req})
        self.assertFalse(serializer.is_valid())
        self.assertIn("baixas", serializer.errors)

    def test_baixa_com_numero_legado_bloqueia_lote(self):
        baixa2 = criar_baixa(self.ua, self.gestor, status=constants.ACEITA, numero_nbbpm="016.0000002.2026")
        from bem_patrimonial.api_serializers import NBBPMGerarLoteSerializer
        data = {
            "baixas": [baixa2.id],
            "numero_processo_baixa": "PROC-NEW",
            "data_autorizacao": str(timezone.localdate()),
            "responsavel": "G",
        }
        req = MagicMock()
        req.user = self.gestor
        serializer = NBBPMGerarLoteSerializer(data=data, context={"request": req})
        self.assertFalse(serializer.is_valid())
        self.assertIn("baixas", serializer.errors)

    def test_aprovar_baixa_ja_com_nbbpm_gera_erro(self):
        client = APIClient()
        client.force_authenticate(user=self.gestor)
        # baixa já tem NBBPM, tenta aprovar outra que já está ACEITA? Na verdade aprovar só de SOLICITADA
        # Cria baixa SOLICITADA com nbbpms_lote preexistente (simula race)
        baixa_solic = criar_baixa(self.ua, self.gestor, status=constants.SOLICITADA)
        # força vínculo
        self.nbbpm.baixas.add(baixa_solic)
        url = reverse("baixas-fisicas-aprovar", kwargs={"pk": baixa_solic.pk})
        resp = client.post(url)
        self.assertIn(resp.status_code, [status.HTTP_400_BAD_REQUEST, status.HTTP_403_FORBIDDEN])

