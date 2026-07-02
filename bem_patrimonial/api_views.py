from rest_framework import viewsets, status, filters, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, CharFilter, DateFilter, ChoiceFilter
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from drf_spectacular.utils import extend_schema, OpenApiResponse
import django_filters

from dados_comuns.models import HistoricoGeral
from django.contrib.contenttypes.models import ContentType

from .models import BaixaFisicaBemPatrimonial, BaixaFisicaBensItem, NBBPM
from .api_serializers import (
    BaixaFisicaBemPatrimonialListSerializer,
    BaixaFisicaBemPatrimonialDetailSerializer,
    BaixaFisicaBemPatrimonialCreateSerializer,
    BaixaFisicaBemPatrimonialUpdateSerializer,
    BaixaFisicaEnviarSolicitacaoSerializer,
    BaixaFisicaAprovarSerializer,
    BaixaFisicaCancelarSerializer,
    BaixaFisicaSolicitarCorrecaoSerializer,
    NBBPMGerarLoteSerializer,
    NBBPMSerializer,
)
from .api_docs import (
    LIST_BAIXAS_FISICAS_DOC,
    CREATE_BAIXAS_FISICAS_DOC,
    RETRIEVE_BAIXAS_FISICAS_DOC,
    UPDATE_BAIXAS_FISICAS_DOC,
    ENVIAR_SOLICITACAO_DOC,
    APROVAR_BAIXA_FISICA_DOC,
    CANCELAR_BAIXA_FISICA_DOC,
    SOLICITAR_CORRECAO_DOC,
    GERAR_NBBPM_DOC,
    EXPORTAR_EXCEL_DOC,
)
from .emails import (
    envia_email_baixa_fisica_solicitada,
    envia_email_baixa_fisica_aprovada,
    envia_email_baixa_fisica_cancelada,
    envia_email_baixa_fisica_correcao_solicitada,
)
from .nbbpm import http_response_nbbpm, gerar_numero_nbbpm
from .nbbpm_lote import (
    http_response_nbbpm_lote,
    gerar_numero_nbbpm_lote,
)
from .laudo_avaliacao import http_response_laudo_avaliacao
from . import constants
from dados_comuns.escopo import filtrar_queryset_por_escopo
from dados_comuns.context import set_user


class IsGestorPatrimonioOrOperadorInventario(IsAuthenticated):
    """
    Permissão customizada: usuário deve ser autenticado e ser
    gestor de patrimônio, operador de inventário ou superuser.
    """

    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        user = request.user
        return (
            user.is_gestor_patrimonio or
            user.is_operador_inventario or
            user.is_superuser
        )


class BaixaFisicaFilter(FilterSet):
    """
    FilterSet customizado para BaixaFisicaBemPatrimonial.

    Suporta:
    - status: filtro exato ou lista (via 'in' usando múltiplos valores separados por vírgula)
    - unidade_administrativa_origem: filtro exato por ID
    - data_aprovacao__gte / data_aprovacao__lte: range de data de aprovação (aceita date string yyyy-MM-dd)
    - data_criacao__gte  / data_criacao__lte:  range de data de criação/solicitação
    """

    status = django_filters.CharFilter(method='filter_status')

    unidade_administrativa_origem = django_filters.NumberFilter(
        field_name='unidade_administrativa_origem__id'
    )

    # Filtros de data de aprovação
    data_aprovacao__gte = django_filters.DateFilter(
        field_name='data_aprovacao',
        lookup_expr='date__gte',
    )
    data_aprovacao__lte = django_filters.DateFilter(
        field_name='data_aprovacao',
        lookup_expr='date__lte',
    )

    # Filtros de data de criação (= data de solicitação)
    data_criacao__gte = django_filters.DateFilter(
        field_name='data_criacao',
        lookup_expr='date__gte',
    )
    data_criacao__lte = django_filters.DateFilter(
        field_name='data_criacao',
        lookup_expr='date__lte',
    )

    class Meta:
        model = BaixaFisicaBemPatrimonial
        fields = []

    def filter_status(self, queryset, name, value):
        """
        Aceita um único valor ('solicitada') ou múltiplos separados por vírgula
        ('solicitada,aguardando_envio'), mapeando para __in quando necessário.
        """
        if not value:
            return queryset
        values = [v.strip() for v in value.split(',') if v.strip()]
        if len(values) == 1:
            return queryset.filter(status=values[0])
        return queryset.filter(status__in=values)


class BaixaFisicaBemPatrimonialViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet,
):
    """
    ViewSet para Baixas Físicas de Bens Patrimoniais.

    Endpoints:
    - GET    /api/baixa-fisica/                         → Listar
    - POST   /api/baixa-fisica/                         → Criar
    - GET    /api/baixa-fisica/{id}/                    → Detalhar
    - PUT    /api/baixa-fisica/{id}/                    → Atualizar
    - PATCH  /api/baixa-fisica/{id}/                    → Atualização parcial
    - POST   /api/baixa-fisica/{id}/aprovar/            → Aprovar
    - POST   /api/baixa-fisica/{id}/recusar/            → Cancelar/Recusar
    - POST   /api/baixa-fisica/{id}/solicitar/          → Enviar para aprovação
    - GET    /api/baixa-fisica/{id}/gerar-nbbpm/        → Gerar PDF NBBPM
    - POST   /api/baixa-fisica/gerar-nbbpm-lote/        → Gerar PDF NBBPM consolidada (NOVO)
    - GET    /api/baixa-fisica/{id}/gerar-laudo/        → Gerar PDF Laudo de Avaliação
    - GET    /api/baixa-fisica/{id}/historico/          → Histórico de alterações
    - GET    /api/baixa-fisica/exportar-excel/          → Exportar Excel
    """

    permission_classes = [IsGestorPatrimonioOrOperadorInventario]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = BaixaFisicaFilter

    search_fields = (
        "numero_processo_baixa",
        "numero_nbbpm",
        "unidade_administrativa_origem__nome",
        "unidade_administrativa_origem__sigla",
        "unidade_administrativa_origem__codigo",
        "criado_por__username",
        "aprovado_por__username",
        "itens__bem__numero_patrimonial",
        "itens__bem__nome",
    )

    ordering_fields = [
        'id',
        'data_criacao',
        'data_aprovacao',
        'status',
        'numero_nbbpm',
        'unidade_administrativa_origem__sigla',
        'numero_processo_baixa',
    ]
    ordering = ['-data_criacao']

    def get_queryset(self):
        queryset = BaixaFisicaBemPatrimonial.objects.all()
        queryset = filtrar_queryset_por_escopo(
            usuario=self.request.user,
            queryset=queryset,
            campo_ua='unidade_administrativa_origem'
        )
        queryset = queryset.select_related(
            'unidade_administrativa_origem',
            'criado_por',
            'aprovado_por'
        ).prefetch_related('itens__bem')
        return queryset.distinct()

    def get_serializer_class(self):
        if self.action == 'list':
            return BaixaFisicaBemPatrimonialListSerializer
        elif self.action == 'retrieve':
            return BaixaFisicaBemPatrimonialDetailSerializer
        elif self.action == 'create':
            return BaixaFisicaBemPatrimonialCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return BaixaFisicaBemPatrimonialUpdateSerializer
        elif self.action == 'enviar_solicitacao':
            return BaixaFisicaEnviarSolicitacaoSerializer
        elif self.action == 'aprovar':
            return BaixaFisicaAprovarSerializer
        elif self.action == 'recusar':
            return BaixaFisicaCancelarSerializer
        elif self.action == 'solicitar_correcao':
            return BaixaFisicaSolicitarCorrecaoSerializer
        elif self.action == 'gerar_nbbpm_lote':
            return NBBPMGerarLoteSerializer
        return BaixaFisicaBemPatrimonialDetailSerializer

    def _detail_response(self, baixa, request, http_status=status.HTTP_200_OK):
        """Retorna resposta com o serializer de detalhe completo."""
        return Response(
            BaixaFisicaBemPatrimonialDetailSerializer(
                baixa, context={'request': request}
            ).data,
            status=http_status,
        )

    # =========================================================
    # HISTÓRICO — UTILITÁRIOS
    # =========================================================

    def _get_content_type(self):
        return ContentType.objects.get_for_model(BaixaFisicaBemPatrimonial)

    def _registrar_historico(self, baixa, campo, valor_antigo, valor_novo, usuario, justificativa=""):
        """Registra uma entrada no histórico geral para a baixa física."""
        HistoricoGeral.objects.create(
            content_type=self._get_content_type(),
            object_id=str(baixa.pk),
            campo=campo,
            valor_antigo=str(valor_antigo) if valor_antigo is not None else "",
            valor_novo=str(valor_novo) if valor_novo is not None else "",
            alterado_por=usuario,
            justificativa=justificativa,
        )

    def _registrar_historico_bulk(self, baixa, registros, usuario):
        """
        Registra múltiplas entradas no histórico geral de uma só vez.
        `registros` é uma lista de dicts com chaves: campo, valor_antigo, valor_novo, justificativa (opcional).
        """
        ct = self._get_content_type()
        HistoricoGeral.objects.bulk_create([
            HistoricoGeral(
                content_type=ct,
                object_id=str(baixa.pk),
                campo=r["campo"],
                valor_antigo=str(r.get("valor_antigo") or ""),
                valor_novo=str(r.get("valor_novo") or ""),
                alterado_por=usuario,
                justificativa=r.get("justificativa", ""),
            )
            for r in registros
        ])

    def _registrar_alteracoes_itens(self, baixa, itens_antes, itens_depois, usuario):
        """
        Compara os bens antes e depois da edição e registra no histórico
        os bens adicionados e os bens removidos.
        """
        ids_antes = set(itens_antes)
        ids_depois = set(itens_depois)

        adicionados = ids_depois - ids_antes
        removidos = ids_antes - ids_depois

        registros = []

        for bem_id in adicionados:
            registros.append({
                "campo": "itens",
                "valor_antigo": "",
                "valor_novo": str(itens_depois[bem_id]),
                "justificativa": "Bem adicionado à baixa",
            })

        for bem_id in removidos:
            registros.append({
                "campo": "itens",
                "valor_antigo": str(itens_antes[bem_id]),
                "valor_novo": "",
                "justificativa": "Bem removido da baixa",
            })

        if registros:
            self._registrar_historico_bulk(baixa, registros, usuario)

    # =========================================================
    # LIST
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Listar baixas físicas",
        description=LIST_BAIXAS_FISICAS_DOC,
        responses={200: BaixaFisicaBemPatrimonialListSerializer(many=True)},
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    # =========================================================
    # CREATE
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Criar baixa física",
        description=CREATE_BAIXAS_FISICAS_DOC,
        responses={
            201: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
        },
    )
    def create(self, request, *args, **kwargs):
        set_user(request.user)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        baixa = serializer.save()

        # Registra criação no histórico
        ct = self._get_content_type()
        registros = [
            HistoricoGeral(
                content_type=ct,
                object_id=str(baixa.pk),
                campo="",
                valor_antigo="",
                valor_novo="",
                alterado_por=request.user,
                justificativa="Baixa física criada",
            )
        ]
        # Registra cada bem incluído na criação
        for item in baixa.itens.select_related('bem'):
            registros.append(HistoricoGeral(
                content_type=ct,
                object_id=str(baixa.pk),
                campo="itens",
                valor_antigo="",
                valor_novo=str(item.bem),
                alterado_por=request.user,
                justificativa="Bem incluído na criação da baixa",
            ))
        HistoricoGeral.objects.bulk_create(registros)

        return self._detail_response(baixa, request, http_status=status.HTTP_201_CREATED)

    # =========================================================
    # RETRIEVE
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Detalhar baixa física",
        description=RETRIEVE_BAIXAS_FISICAS_DOC,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    # =========================================================
    # UPDATE / PARTIAL UPDATE
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Atualizar baixa física",
        description=UPDATE_BAIXAS_FISICAS_DOC,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    def update(self, request, *args, **kwargs):
        set_user(request.user)
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # Captura estado dos itens ANTES da atualização
        itens_antes = {
            item.bem_id: str(item.bem)
            for item in instance.itens.select_related('bem')
        }

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        baixa = self.get_queryset().get(pk=instance.pk)

        # Captura estado dos itens DEPOIS e registra diferenças
        itens_depois = {
            item.bem_id: str(item.bem)
            for item in baixa.itens.select_related('bem')
        }
        self._registrar_alteracoes_itens(baixa, itens_antes, itens_depois, request.user)

        return self._detail_response(baixa, request)

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Atualização parcial",
        description=UPDATE_BAIXAS_FISICAS_DOC,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
        },
    )
    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    # =========================================================
    # HISTÓRICO
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Histórico de alterações",
        description=(
            "Retorna o histórico de alterações da baixa física, "
            "seguindo o padrão dos demais módulos da API."
        ),
        request=None,
        responses={
            200: OpenApiResponse(description="Lista de registros do histórico"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=["get"], url_path="historico")
    def historico(self, request, pk=None):
        instance = self.get_object()

        ct = ContentType.objects.get_for_model(BaixaFisicaBemPatrimonial)

        registros = (
            HistoricoGeral.objects
            .filter(content_type=ct, object_id=str(instance.pk))
            .select_related("alterado_por")
            .order_by("-id")
        )

        data = [
            {
                "id": r.id,
                "campo": r.campo,
                "valor_antigo": r.valor_antigo,
                "valor_novo": r.valor_novo,
                "justificativa": r.justificativa,
                "alterado_por": (
                    r.alterado_por.username if r.alterado_por else None
                ),
                "data_alteracao": r.alterado_em if hasattr(r, "alterado_em") else None,
            }
            for r in registros
        ]

        return Response(data, status=status.HTTP_200_OK)

    # =========================================================
    # ENVIAR SOLICITAÇÃO
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Enviar solicitação",
        description=ENVIAR_SOLICITACAO_DOC,
        request=None,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['post'], url_path='solicitar')
    def enviar_solicitacao(self, request, pk=None):
        baixa = self.get_object()

        serializer = self.get_serializer(
            data={}, context={'baixa': baixa, 'request': request}
        )
        serializer.is_valid(raise_exception=True)

        status_anterior = baixa.get_status_display()

        set_user(request.user)
        baixa.enviar_solicitacao()

        self._registrar_historico(
            baixa=baixa,
            campo="status",
            valor_antigo=status_anterior,
            valor_novo=baixa.get_status_display(),
            usuario=request.user,
            justificativa="Solicitação de baixa enviada para aprovação",
        )

        try:
            envia_email_baixa_fisica_solicitada(baixa)
        except Exception:
            pass

        return self._detail_response(baixa, request)

    # =========================================================
    # APROVAR
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Aprovar baixa física",
        description=APROVAR_BAIXA_FISICA_DOC,
        request=None,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
            403: OpenApiResponse(description="Sem permissão"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['post'])
    def aprovar(self, request, pk=None):
        baixa = self.get_object()

        serializer = self.get_serializer(
            data={}, context={'baixa': baixa, 'request': request}
        )
        serializer.is_valid(raise_exception=True)

        status_anterior = baixa.get_status_display()
        nbbpm_anterior = baixa.numero_nbbpm

        with transaction.atomic():
            set_user(request.user)
            baixa.aprovar(usuario_aprovador=request.user)

            nbbpm_gerado = None
            if not baixa.numero_nbbpm:
                nbbpm_gerado = gerar_numero_nbbpm(baixa)
                baixa.numero_nbbpm = nbbpm_gerado
                baixa.save(update_fields=['numero_nbbpm'])

        # Registra mudança de status
        self._registrar_historico(
            baixa=baixa,
            campo="status",
            valor_antigo=status_anterior,
            valor_novo=baixa.get_status_display(),
            usuario=request.user,
            justificativa="Baixa física aprovada",
        )

        # Registra geração do NBBPM se ocorreu
        if nbbpm_gerado:
            self._registrar_historico(
                baixa=baixa,
                campo="numero_nbbpm",
                valor_antigo=nbbpm_anterior or "",
                valor_novo=nbbpm_gerado,
                usuario=request.user,
                justificativa="Número NBBPM gerado automaticamente na aprovação",
            )

        try:
            envia_email_baixa_fisica_aprovada(baixa)
        except Exception:
            pass

        return self._detail_response(baixa, request)

    # =========================================================
    # CANCELAR / RECUSAR
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Recusar baixa física",
        description=CANCELAR_BAIXA_FISICA_DOC,
        request=BaixaFisicaCancelarSerializer,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
            403: OpenApiResponse(description="Sem permissão"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['post'], url_path='recusar')
    def recusar(self, request, pk=None):
        baixa = self.get_object()

        serializer = self.get_serializer(
            data=request.data, context={'baixa': baixa, 'request': request}
        )
        serializer.is_valid(raise_exception=True)

        status_anterior = baixa.get_status_display()
        motivo = serializer.validated_data.get('motivo', '')

        with transaction.atomic():
            set_user(request.user)
            self._restaurar_bens_da_baixa(baixa)
            baixa.status = constants.RECUSADA
            baixa.save(update_fields=['status'])

        self._registrar_historico(
            baixa=baixa,
            campo="status",
            valor_antigo=status_anterior,
            valor_novo=baixa.get_status_display(),
            usuario=request.user,
            justificativa=f"Baixa recusada. Motivo: {motivo}" if motivo else "Baixa recusada",
        )

        try:
            envia_email_baixa_fisica_cancelada(baixa, request.user)
        except Exception:
            pass

        return self._detail_response(baixa, request)

    # =========================================================
    # SOLICITAR CORREÇÃO
    # =========================================================
    #
    # Distinto de `recusar`: este fluxo é uma devolução para ajustes,
    # não um encerramento do processo. Usado pela tela "Validar Baixa"
    # do frontend, através da página própria "Solicitar correção".
    #
    # Diferenças em relação a `recusar`:
    #   - só é permitido a partir do status "solicitada"
    #   - o status final é "aguardando_envio" (Em elaboração), não
    #     "recusada"
    #   - motivo é obrigatório, não opcional
    #   - os bens NÃO são restaurados ao status "aprovado": continuam
    #     em "Baixa Física - Aguardando aprovação", pois a baixa ainda
    #     está em andamento, apenas voltando para edição

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Solicitar correção da baixa física",
        description=SOLICITAR_CORRECAO_DOC,
        request=BaixaFisicaSolicitarCorrecaoSerializer,
        responses={
            200: BaixaFisicaBemPatrimonialDetailSerializer,
            400: OpenApiResponse(description="Erro de validação"),
            403: OpenApiResponse(description="Sem permissão"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['post'], url_path='solicitar-correcao')
    def solicitar_correcao(self, request, pk=None):
        baixa = self.get_object()

        serializer = self.get_serializer(
            data=request.data, context={'baixa': baixa, 'request': request}
        )
        serializer.is_valid(raise_exception=True)

        status_anterior = baixa.get_status_display()
        motivo = serializer.validated_data['motivo']

        with transaction.atomic():
            set_user(request.user)
            baixa.status = constants.AGUARDANDO_ENVIO
            baixa.save(update_fields=['status'])

        self._registrar_historico(
            baixa=baixa,
            campo="status",
            valor_antigo=status_anterior,
            valor_novo=baixa.get_status_display(),
            usuario=request.user,
            justificativa=f"Correção solicitada. Observações: {motivo}",
        )

        try:
            envia_email_baixa_fisica_correcao_solicitada(baixa, request.user)
        except Exception:
            pass

        return self._detail_response(baixa, request)

    def _restaurar_bens_da_baixa(self, baixa: BaixaFisicaBemPatrimonial) -> None:
        """Restaura o status dos bens ao cancelar uma baixa."""
        for item in baixa.itens.select_related('bem'):
            bem = item.bem
            if bem.status == constants.BAIXA_FISICA_AGUARDANDO_APROVACAO:
                bem.status = constants.APROVADO
                bem.save(update_fields=['status'])

    # =========================================================
    # GERAR NBBPM
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Gerar PDF NBBPM",
        description=GERAR_NBBPM_DOC,
        request=None,
        responses={
            200: OpenApiResponse(description="PDF da Nota NBBPM"),
            400: OpenApiResponse(description="Baixa não aprovada ou sem NBBPM"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['get'], url_path='gerar-nbbpm')
    def gerar_nbbpm(self, request, pk=None):
        baixa = self.get_object()

        if baixa.status != constants.ACEITA:
            return Response(
                {'detail': 'A Nota NBBPM só pode ser gerada para baixas aprovadas.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not baixa.numero_nbbpm:
            return Response(
                {'detail': 'Esta baixa não possui número NBBPM.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return http_response_nbbpm(baixa)

    # =========================================================
    # GERAR NBBPM CONSOLIDADA (LOTE) — NOVO
    # =========================================================
    #
    # Fluxo da tela "Baixas Físicas de Bens Patrimoniais": o usuário
    # seleciona uma ou mais Baixas com status Aprovado (ACEITA) da sua
    # Unidade Orçamentária, informa Número do processo de Baixa, Data da
    # Autorização, Responsável e (opcionalmente) o Número do processo de
    # destinação final, e o sistema gera um único PDF consolidado.

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Gerar PDF NBBPM consolidada (lote)",
        description=(
            "Gera uma NBBPM consolidada a partir de uma ou mais Baixas "
            "Físicas aprovadas (status ACEITA) da mesma Unidade "
            "Orçamentária do usuário logado."
        ),
        request=NBBPMGerarLoteSerializer,
        responses={
            200: OpenApiResponse(description="PDF da NBBPM consolidada"),
            400: OpenApiResponse(description="Seleção inválida de Baixas Físicas"),
        },
    )
    @action(detail=False, methods=['post'], url_path='gerar-nbbpm-lote')
    def gerar_nbbpm_lote(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            nbbpm = serializer.save()
            nbbpm.numero = gerar_numero_nbbpm_lote(nbbpm)
            nbbpm.save(update_fields=['numero'])

        return http_response_nbbpm_lote(nbbpm, usuario_gerador=request.user)

    # =========================================================
    # GERAR LAUDO DE AVALIAÇÃO
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Gerar PDF do Laudo de Avaliação",
        description=(
            "Gera o Laudo de Avaliação para Baixa de Bens Patrimoniais Móveis "
            "conforme Artigo 20 do Decreto 53.484/2012. "
            "Disponível apenas para baixas com status 'Aceita'."
        ),
        request=None,
        responses={
            200: OpenApiResponse(description="PDF do Laudo de Avaliação"),
            400: OpenApiResponse(description="Baixa não está com status Aceita"),
            404: OpenApiResponse(description="Baixa física não encontrada"),
        },
    )
    @action(detail=True, methods=['get'], url_path='gerar-laudo')
    def gerar_laudo(self, request, pk=None):
        baixa = self.get_object()

        if baixa.status != constants.ACEITA:
            return Response(
                {'detail': 'O Laudo de Avaliação só pode ser gerado para baixas aceitas.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return http_response_laudo_avaliacao(baixa)

    # =========================================================
    # EXPORTAR EXCEL
    # =========================================================

    @extend_schema(
        tags=["Baixas Físicas"],
        summary="Exportar para Excel",
        description=EXPORTAR_EXCEL_DOC,
        request=None,
        responses={200: OpenApiResponse(description="Arquivo Excel")},
    )
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        queryset = self._get_queryset_para_exportacao(request)
        wb = self._construir_workbook(queryset)
        return self._montar_resposta_excel(wb)

    def _get_queryset_para_exportacao(self, request):
        """Aplica filtros e restrição por IDs ao queryset de exportação."""
        queryset = self.filter_queryset(self.get_queryset())
        ids_param = request.query_params.get('ids')
        if ids_param:
            ids_list = self._parse_ids(ids_param)
            if ids_list:
                queryset = queryset.filter(id__in=ids_list)
        return queryset

    def _parse_ids(self, ids_param: str) -> list:
        """Converte string de IDs separados por vírgula em lista de ints."""
        ids = []
        for id_str in ids_param.split(','):
            try:
                ids.append(int(id_str.strip()))
            except (ValueError, TypeError):
                pass
        return ids

    def _construir_workbook(self, queryset) -> Workbook:
        """Monta o workbook Excel com cabeçalho e dados."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Baixas Físicas"
        self._escrever_cabecalho(ws)
        self._escrever_dados(ws, queryset)
        self._ajustar_colunas(ws)
        return wb

    def _escrever_cabecalho(self, ws) -> None:
        headers = [
            'Unidade Administrativa', 'Número Patrimonial', 'Nome do Bem',
            'Status', 'NBBPM', 'Usuário que solicitou a Baixa',
            'Gestor que aprovou a solicitação', 'Data da Aprovação',
        ]
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _escrever_dados(self, ws, queryset) -> None:
        row_num = 2
        for baixa in queryset:
            itens_baixa = list(baixa.itens.all())
            if not itens_baixa:
                self._escrever_linha(ws, row_num, baixa, bem=None)
                row_num += 1
            else:
                for item in itens_baixa:
                    self._escrever_linha(ws, row_num, baixa, bem=item.bem)
                    row_num += 1

        cell_alignment = Alignment(vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = cell_alignment

    def _escrever_linha(self, ws, row_num: int, baixa, bem) -> None:
        data_aprovacao = (
            baixa.data_aprovacao.strftime('%d/%m/%Y %H:%M')
            if baixa.data_aprovacao else '-'
        )
        ws.cell(row=row_num, column=1).value = (
            str(baixa.unidade_administrativa_origem)
            if baixa.unidade_administrativa_origem else '-'
        )
        ws.cell(row=row_num, column=2).value = bem.numero_patrimonial if bem else '-'
        ws.cell(row=row_num, column=3).value = bem.nome if bem else '-'
        ws.cell(row=row_num, column=4).value = baixa.get_status_display()
        ws.cell(row=row_num, column=5).value = baixa.numero_nbbpm or '-'
        ws.cell(row=row_num, column=6).value = str(baixa.criado_por) if baixa.criado_por else '-'
        ws.cell(row=row_num, column=7).value = str(baixa.aprovado_por) if baixa.aprovado_por else '-'
        ws.cell(row=row_num, column=8).value = data_aprovacao

    def _ajustar_colunas(self, ws) -> None:
        for i, width in enumerate([30, 20, 35, 20, 20, 30, 30, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _montar_resposta_excel(self, wb: Workbook) -> HttpResponse:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'baixas_fisicas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
