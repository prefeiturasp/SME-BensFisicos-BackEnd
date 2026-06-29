import io
from dados_comuns.tests.auth_test_utils import (
    NEW_PASSWORD1_KEY,
    NEW_PASSWORD2_KEY,
    auth_kwargs,
)
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.conf import settings
from django.utils import timezone
import uuid
from rest_framework.test import APIClient, APITestCase
from openpyxl import load_workbook

from dados_comuns.tests.factories import criar_ua, criar_uo
from usuario.constants import GRUPO_GESTOR_PATRIMONIO, GRUPO_OPERADOR_INVENTARIO

User = get_user_model()


class PasswordChangeViewTests(TestCase):

    def setUp(self):
        self.user = User.objects.create_user(
            username="alice", **auth_kwargs("a123456"), must_change_password=False
        )
        self.staff = User.objects.create_user(
            username="admin1",
            **auth_kwargs("a123456"),
            is_staff=True,
            must_change_password=False,
        )

    def test_get_own_password_change_page(self):
        self.client.login(username="alice", **auth_kwargs("a123456"))
        url = reverse("password_change")

        resp = self.client.get(url)

        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, "admin/password_change.html")

    def test_post_own_password_change_sets_flags_and_redirects_next(self):
        self.client.login(username="alice", **auth_kwargs("a123456"))
        next_url = "/admin/"

        url = f'{reverse("password_change")}?next={next_url}'

        resp = self.client.post(
            url,
            {
                NEW_PASSWORD1_KEY: "N0va@s3nhA!",
                NEW_PASSWORD2_KEY: "N0va@s3nhA!",
                "next": next_url,
            },
            follow=False,
        )

        self.assertEqual(resp.status_code, 302)
        self.assertTrue(resp["Location"].endswith(next_url))

        u = User.objects.get(pk=self.user.pk)

        self.assertFalse(u.must_change_password)
        self.assertIsNotNone(u.last_password_change)
        self.assertLessEqual(u.last_password_change, timezone.now())
        self.assertTrue(self.client.login(username="alice", **auth_kwargs("N0va@s3nhA!")))

    def test_staff_changes_other_user_password_without_old_password(self):
        self.client.login(username="admin1", **auth_kwargs("a123456"))
        target = self.user

        next_url = f"/admin/usuario/usuario/{target.pk}/change/"

        url = f'{reverse("password_change")}?user_id={target.pk}&next={next_url}'

        resp = self.client.post(
            url,
            {
                NEW_PASSWORD1_KEY: "Sup3rS3nh@",
                NEW_PASSWORD2_KEY: "Sup3rS3nh@",
                "user_id": str(target.pk),
                "next": next_url,
            },
            follow=False,
        )

        self.assertEqual(resp.status_code, 302)

        self.assertTrue(resp["Location"].endswith(next_url))

        target.refresh_from_db()

        self.assertFalse(target.must_change_password)
        self.assertIsNotNone(target.last_password_change)
        self.assertTrue(self.client.login(username="alice", **auth_kwargs("Sup3rS3nh@")))

    def test_redirecionamento_admin_senha_exige_autenticacao(self):
        url = reverse("admin_usuario_password_redirect", args=[self.user.pk])
        resp = self.client.get(url)

        self.assertEqual(resp.status_code, 302)
        self.assertIn("/admin/login/", resp["Location"])
        self.assertIn("next=", resp["Location"])

    def test_redirecionamento_admin_senha_nega_usuario_nao_staff(self):
        self.client.login(username="alice", **auth_kwargs("a123456"))
        url = reverse("admin_usuario_password_redirect", args=[self.user.pk])
        resp = self.client.get(url)

        self.assertEqual(resp.status_code, 403)

    def test_redirecionamento_admin_senha_permite_staff_via_get(self):
        self.client.login(username="admin1", **auth_kwargs("a123456"))
        url = reverse("admin_usuario_password_redirect", args=[self.user.pk])
        resp = self.client.get(url)

        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("password_change"), resp["Location"])
        self.assertIn(f"user_id={self.user.pk}", resp["Location"])

    def test_redirecionamento_admin_senha_rejeita_post(self):
        self.client.login(username="admin1", **auth_kwargs("a123456"))
        url = reverse("admin_usuario_password_redirect", args=[self.user.pk])
        resp = self.client.post(url)

        self.assertEqual(resp.status_code, 405)


class AdminLoginViewTests(TestCase):

    def setUp(self):
        self.senha_list = ["S3nh", "@", "123"]
        self.senha = "".join(self.senha_list)
        self.user = User.objects.create_user(username="bob", password=self.senha)

    def test_login_redirects_selecionar_ua(self):

        resp = self.client.post(
            "/admin/login/",
            {
                "username": "bob",
                "password": self.senha,
            },
            follow=False,
        )

        self.assertEqual(resp.status_code, 302)

        self.assertEqual(resp["Location"], reverse("selecionar_ua"))

    def test_login_redirects_selecionar_ua_preserving_next(self):

        resp = self.client.post(
            "/admin/login/",
            {
                "username": "bob",
                "password": self.senha,
                "next": "/admin/",
            },
            follow=False,
        )

        self.assertEqual(resp.status_code, 302)

        self.assertIn(reverse("selecionar_ua"), resp["Location"])
        self.assertIn("next=%2Fadmin%2F", resp["Location"])


class AdminGoogleAnalyticsTests(TestCase):

    def setUp(self):
        self.staff_user = User.objects.create_user(
            username="analytics_admin",
            **auth_kwargs("analytics-admin-123"),
            is_staff=True,
            is_superuser=True,
            must_change_password=False,
        )

    def test_admin_login_exibe_google_analytics_em_producao(self):
        with self.settings(DEBUG=False):
            response = self.client.get(reverse("admin:login"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f"https://www.googletagmanager.com/gtag/js?id={settings.GOOGLE_ANALYTICS_ID}",
        )
        self.assertContains(
            response,
            f"gtag('config', '{settings.GOOGLE_ANALYTICS_ID}');",
            html=False,
        )

    def test_admin_login_nao_exibe_google_analytics_em_desenvolvimento(self):
        with self.settings(DEBUG=True):
            response = self.client.get(reverse("admin:login"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "www.googletagmanager.com/gtag/js")
        self.assertNotContains(response, "gtag('config'")

    def test_admin_index_exibe_google_analytics_em_producao(self):
        self.client.force_login(self.staff_user)

        with self.settings(DEBUG=False):
            response = self.client.get(reverse("admin:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f"https://www.googletagmanager.com/gtag/js?id={settings.GOOGLE_ANALYTICS_ID}",
        )
        self.assertContains(
            response,
            f"gtag('config', '{settings.GOOGLE_ANALYTICS_ID}');",
            html=False,
        )


class SelecionarUAViewTests(TestCase):

    def setUp(self):

        self.senha_list = ["S3nh", "@", "123"]
        self.senha = "".join(self.senha_list)

        self.ua1 = criar_ua(codigo="001", sigla="UA1", nome="Unidade 1")

        self.ua2 = criar_ua(
            uo=self.ua1.unidade_orcamentaria,
            codigo="002",
            sigla="UA2",
            nome="Unidade 2",
        )

        self.grupo_gestor = Group.objects.get_or_create(
            name=GRUPO_GESTOR_PATRIMONIO
        )[0]

        self.grupo_operador = Group.objects.get_or_create(
            name=GRUPO_OPERADOR_INVENTARIO
        )[0]

        self.gestor = User.objects.create_user(
            username="gestor1",
            password=self.senha,
            unidade_orcamentaria=self.ua1.unidade_orcamentaria,
            unidade_administrativa=self.ua1,
            is_staff=True,
            must_change_password=False,
        )

        self.gestor.groups.add(self.grupo_gestor)

        self.operador = User.objects.create_user(
            username="operador1",
            password=self.senha,
            unidade_orcamentaria=self.ua1.unidade_orcamentaria,
            unidade_administrativa=self.ua1,
            is_staff=True,
            must_change_password=False,
        )

        self.operador.groups.add(self.grupo_operador)

        self.operador.unidades_administrativas.add(self.ua1, self.ua2)

    def test_gestor_ve_opcao_visao_geral_no_select(self):

        self.client.login(username="gestor1", password=self.senha)

        resp = self.client.get(reverse("selecionar_ua"))

        self.assertEqual(resp.status_code, 200)

        self.assertContains(resp, "visão geral")
        self.assertContains(resp, "__UO__")

    def test_gestor_pode_selecionar_visao_geral(self):

        self.client.login(username="gestor1", password=self.senha)

        resp = self.client.post(
            reverse("selecionar_ua"),
            {
                "unidade_administrativa": "__UO__",
                "next": "/admin/",
            },
            follow=False,
        )

        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp["Location"], "/admin/")

        self.gestor.refresh_from_db()

        self.assertIsNone(self.gestor.unidade_administrativa_id)

        self.assertEqual(
            self.gestor.unidade_orcamentaria_id,
            self.ua1.unidade_orcamentaria_id,
        )

    def test_operador_nao_ve_opcao_visao_geral(self):

        self.client.login(username="operador1", password=self.senha)

        resp = self.client.get(reverse("selecionar_ua"))

        self.assertEqual(resp.status_code, 200)

        self.assertNotContains(resp, "__UO__")


class UsuarioViewSetTests(TestCase):

    def setUp(self):

        self.senha_list = ["S3nh", "@", "123"]
        self.senha = "".join(self.senha_list)

        self.client = APIClient()

        self.admin = User.objects.create_user(
            username="admin",
            password=self.senha,
            is_staff=True,
            is_superuser=True,
        )

        self.user = User.objects.create_user(
            username="user1",
            email="user@test.com",
            password=self.senha,
        )

        self.client.login(username="admin", password=self.senha)

        self.client.force_authenticate(user=self.admin)

        self.list_url = reverse("usuario-list")

        self.detail_url = reverse("usuario-detail", args=[self.user.id])

        self.restore_url = reverse("usuario-restore", args=[self.user.id])

    def test_list_users(self):

        resp = self.client.get(self.list_url)

        self.assertEqual(resp.status_code, 200)

        self.assertIn("results", resp.data)

    def test_retrieve_user(self):

        resp = self.client.get(self.detail_url)

        self.assertEqual(resp.status_code, 200)

        self.assertEqual(resp.data["username"], "user1")

    def test_create_user(self):

        resp = self.client.post(
            self.list_url,
            {
                "username": "novo",
                "email": "novo@test.com",
                "password": self.senha,
                "password_confirm": self.senha,
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 201)

        self.assertTrue(User.objects.filter(username="novo").exists())

        user = User.objects.get(username="novo")

        self.assertTrue(user.must_change_password)
        self.assertIsNotNone(user.last_password_change)

    def test_update_user(self):

        resp = self.client.put(
            self.detail_url,
            {
                "username": "user1",
                "email": "novo@email.com",
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 200)

        self.user.refresh_from_db()

        self.assertEqual(self.user.email, "novo@email.com")

    def test_partial_update_user(self):

        resp = self.client.patch(
            self.detail_url,
            {
                "email": "patch@email.com",
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 200)

        self.user.refresh_from_db()

        self.assertEqual(self.user.email, "patch@email.com")

    def test_soft_delete_user(self):

        resp = self.client.delete(self.detail_url)

        self.assertEqual(resp.status_code, 200)

        self.user.refresh_from_db()

        self.assertFalse(self.user.is_active)

    def test_restore_user(self):

        self.user.is_active = False
        self.user.save()

        resp = self.client.post(self.restore_url)

        self.assertEqual(resp.status_code, 200)

        self.user.refresh_from_db()

        self.assertTrue(self.user.is_active)

    def test_filter_active_users(self):

        self.user.is_active = False
        self.user.save()

        resp = self.client.get(self.list_url, {"is_active": False})

        self.assertEqual(resp.status_code, 200)

        usernames = [u["username"] for u in resp.data["results"]]

        self.assertIn("user1", usernames)


class UsuarioExportViewSetTests(TestCase):

    def setUp(self):
        self.client = APIClient()
        self.senha = "S3nh@123"

        self.group_gestor = Group.objects.get_or_create(
            name=GRUPO_GESTOR_PATRIMONIO
        )[0]
        self.group_operador = Group.objects.get_or_create(
            name=GRUPO_OPERADOR_INVENTARIO
        )[0]

        self.uo1 = criar_uo(codigo="101", nome="UO 101")
        self.ua1 = criar_ua(
            uo=self.uo1,
            codigo="101.001",
            sigla="UA-101",
            nome="Unidade Administrativa 101",
        )
        self.uo2 = criar_uo(codigo="202", nome="UO 202")
        self.ua2 = criar_ua(
            uo=self.uo2,
            codigo="202.001",
            sigla="UA-202",
            nome="Unidade Administrativa 202",
        )

        self.superuser = User.objects.create_user(
            username="superuser_export",
            email="superuser_export@test.com",
            password=self.senha,
            nome="Superuser Export",
            rf="900001",
            is_staff=True,
            is_superuser=True,
            unidade_orcamentaria=self.uo1,
            unidade_administrativa=self.ua1,
        )
        self.superuser.groups.add(self.group_gestor)

        self.gestor = User.objects.create_user(
            username="gestor_export",
            email="gestor_export@test.com",
            password=self.senha,
            nome="Gestor Export",
            rf="900002",
            is_staff=True,
            unidade_orcamentaria=self.uo1,
            unidade_administrativa=self.ua1,
        )
        self.gestor.groups.add(self.group_gestor)

        self.operador = User.objects.create_user(
            username="operador_export",
            email="operador_export@test.com",
            password=self.senha,
            nome="Operador Export",
            rf="900003",
            is_staff=True,
            unidade_orcamentaria=self.uo1,
            unidade_administrativa=self.ua1,
        )
        self.operador.groups.add(self.group_operador)

        self.usuario_outra_uo = User.objects.create_user(
            username="usuario_outra_uo",
            email="outra@test.com",
            password=self.senha,
            nome="Usuario Outra UO",
            rf="900004",
            is_staff=True,
            unidade_orcamentaria=self.uo2,
            unidade_administrativa=self.ua2,
        )

        self.export_url = reverse("usuario-exportar")

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def _ler_planilha(self, response):
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))

    def test_superuser_exporta_todos_os_usuarios(self):
        self._auth(self.superuser)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])

        rows = self._ler_planilha(response)
        self.assertEqual(
            rows[0],
            ("Nome", "RF", "E-mail", "Unidade Administrativa"),
        )
        nomes = {row[0] for row in rows[1:] if row and row[0]}
        self.assertIn("Gestor Export", nomes)
        self.assertIn("Operador Export", nomes)
        self.assertIn("Usuario Outra UO", nomes)

    def test_gestor_exporta_apenas_usuarios_da_sua_uo(self):
        self._auth(self.gestor)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)

        rows = self._ler_planilha(response)
        nomes = {row[0] for row in rows[1:] if row and row[0]}
        self.assertIn("Gestor Export", nomes)
        self.assertIn("Operador Export", nomes)
        self.assertNotIn("Usuario Outra UO", nomes)

    def test_operador_nao_pode_exportar_usuarios(self):
        self._auth(self.operador)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 403)


# =========================================
# TESTES NEGATIVOS DE SEGURANÇA (AUDITORIA)
# =========================================

class UsuarioPermissionNegativeTests(APITestCase):

    def setUp(self):

        self.senha_list = ["S3nh", "@", "123"]
        self.senha = "".join(self.senha_list)

        self.client = APIClient()

        self.grupo_gestor = Group.objects.get_or_create(
            name=GRUPO_GESTOR_PATRIMONIO
        )[0]

        self.grupo_operador = Group.objects.get_or_create(
            name=GRUPO_OPERADOR_INVENTARIO
        )[0]

        self.ua1 = criar_ua(codigo="001", sigla="UA1", nome="Unidade 1")

        self.ua2 = criar_ua(
            uo=self.ua1.unidade_orcamentaria,
            codigo="002",
            sigla="UA2",
            nome="Unidade 2",
        )

        self.gestor = User.objects.create_user(
            username="gestor",
            password=self.senha,
            unidade_orcamentaria=self.ua1.unidade_orcamentaria,
            unidade_administrativa=self.ua1,
            is_staff=True,
        )

        self.gestor.groups.add(self.grupo_gestor)

        self.operador = User.objects.create_user(
            username="operador",
            password=self.senha,
            unidade_orcamentaria=self.ua1.unidade_orcamentaria,
            unidade_administrativa=self.ua1,
            is_staff=True,
        )

        self.operador.groups.add(self.grupo_operador)

        self.target_user = User.objects.create_user(
            username="target",
            password=self.senha,
            unidade_orcamentaria=self.ua1.unidade_orcamentaria,
            unidade_administrativa=self.ua1,
        )

        self.list_url = reverse("usuario-list")
        self.detail_url = reverse("usuario-detail", args=[self.target_user.id])

    def test_operador_nao_pode_criar_usuario(self):

        self.client.force_authenticate(user=self.operador)

        resp = self.client.post(
            self.list_url,
            {
                "username": "hack",
                "password": self.senha,
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 403)

    def test_operador_nao_pode_editar_usuario(self):

        self.client.force_authenticate(user=self.operador)

        resp = self.client.patch(
            self.detail_url,
            {
                "email": "hack@test.com"
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 403)

    def test_nao_permitir_elevar_para_superuser(self):

        self.client.force_authenticate(user=self.operador)

        resp = self.client.patch(
            self.detail_url,
            {
                "is_superuser": True
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 403)

    def test_nao_permitir_elevar_para_staff(self):

        self.client.force_authenticate(user=self.operador)

        resp = self.client.patch(
            self.detail_url,
            {
                "is_staff": True
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 403)

    def test_nao_permitir_alterar_grupos(self):

        gestor_group = Group.objects.get(name=GRUPO_GESTOR_PATRIMONIO)

        self.client.force_authenticate(user=self.operador)

        resp = self.client.patch(
            self.detail_url,
            {
                "groups": [gestor_group.id]
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 403)

    def test_acesso_fora_do_escopo_uo(self):

        uo_externa = criar_uo(
            codigo="998",
            sigla="UOEXT",
            nome="UO Externa"
        )

        ua_externa = criar_ua(
            codigo="999",
            sigla="EXT",
            nome="Externa",
            uo=uo_externa
        )

        user_externo = User.objects.create_user(
            username="externo",
            password=self.senha,
            unidade_orcamentaria=uo_externa,
            unidade_administrativa=ua_externa,
        )

        self.client.force_authenticate(user=self.operador)

        url = reverse("usuario-detail", args=[user_externo.id])

        resp = self.client.get(url)

        self.assertIn(resp.status_code, [403, 404])

    def test_gestor_nao_pode_definir_superuser(self):

        self.client.force_authenticate(user=self.gestor)

        resp = self.client.patch(
            self.detail_url,
            {
                "is_superuser": True
            },
            format="json",
        )

        self.assertIn(resp.status_code, [400, 403])

    def test_gestor_nao_pode_criar_superuser(self):

        self.client.force_authenticate(user=self.gestor)

        resp = self.client.post(
            self.list_url,
            {
                "username": "superhack",
                "password": self.senha,
                "password_confirm": self.senha,
                "is_superuser": True,
            },
            format="json",
        )

        self.assertIn(resp.status_code, [400, 403])

    def test_gestor_nao_pode_definir_staff(self):

        self.client.force_authenticate(user=self.gestor)

        resp = self.client.patch(
            self.detail_url,
            {
                "is_staff": True
            },
            format="json",
        )

        self.assertIn(resp.status_code, [400, 403])

    def test_usuario_nao_pode_alterar_proprio_grupo(self):

        self.client.force_authenticate(user=self.gestor)

        gestor_group = Group.objects.get(name=GRUPO_GESTOR_PATRIMONIO)

        url = reverse("usuario-detail", args=[self.gestor.id])

        resp = self.client.patch(
            url,
            {
                "groups": [gestor_group.id]
            },
            format="json",
        )

        self.assertIn(resp.status_code, [400, 403])


class UsuarioCreateUoScopeValidationTests(APITestCase):

    def setUp(self):
        self.client = APIClient()
        self.senha = "T3st@123"

        self.grupo_gestor = Group.objects.get_or_create(
            name=GRUPO_GESTOR_PATRIMONIO
        )[0]
        self.grupo_operador = Group.objects.get_or_create(
            name=GRUPO_OPERADOR_INVENTARIO
        )[0]

        self.uo1 = criar_uo(codigo="101", sigla="UO1", nome="UO 1")
        self.uo2 = criar_uo(codigo="102", sigla="UO2", nome="UO 2")
        self.ua1_uo1 = criar_ua(uo=self.uo1, codigo="001", sigla="A1", nome="UA 1/UO1")
        self.ua2_uo1 = criar_ua(uo=self.uo1, codigo="002", sigla="A2", nome="UA 2/UO1")
        self.ua1_uo2 = criar_ua(uo=self.uo2, codigo="003", sigla="B1", nome="UA 1/UO2")

        self.gestor_uo1 = User.objects.create_user(
            username="gestor_uo1",
            password=self.senha,
            unidade_orcamentaria=self.uo1,
            unidade_administrativa=self.ua1_uo1,
            is_staff=True,
        )
        self.gestor_uo1.groups.add(self.grupo_gestor)

        self.list_url = reverse("usuario-list")

    def _payload_base(self):
        return {
            "username": f"op_{uuid.uuid4().hex[:8]}",
            "nome": "Operador Teste",
            "rf": "A12345",
            "email": f"{uuid.uuid4().hex[:8]}@teste.com",
            "password": self.senha,
            "password_confirm": self.senha,
            "group_name": GRUPO_OPERADOR_INVENTARIO,
            "is_active": True,
        }

    def test_create_sem_unidade_orcamentaria_retorna_erro(self):
        self.client.force_authenticate(user=self.gestor_uo1)
        payload = self._payload_base()
        payload.update(
            {
                "unidade_administrativa": self.ua1_uo1.id,
                "unidades_administrativas": [self.ua1_uo1.id],
            }
        )

        response = self.client.post(self.list_url, payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("unidade_orcamentaria", response.data)

    def test_create_operador_com_unidades_vazias_retorna_erro(self):
        self.client.force_authenticate(user=self.gestor_uo1)
        payload = self._payload_base()
        payload.update(
            {
                "unidade_orcamentaria": self.uo1.id,
                "unidade_administrativa": self.ua1_uo1.id,
                "unidades_administrativas": [],
            }
        )

        response = self.client.post(self.list_url, payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("unidades_administrativas", response.data)

    def test_create_com_ua_fora_do_escopo_retorna_erro(self):
        self.client.force_authenticate(user=self.gestor_uo1)
        payload = self._payload_base()
        payload.update(
            {
                "unidade_orcamentaria": self.uo1.id,
                "unidade_administrativa": self.ua1_uo1.id,
                "unidades_administrativas": [self.ua1_uo2.id],
            }
        )

        response = self.client.post(self.list_url, payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("unidades_administrativas", response.data)

    def test_create_valido_no_escopo_do_gestor(self):
        self.client.force_authenticate(user=self.gestor_uo1)
        payload = self._payload_base()
        payload.update(
            {
                "unidade_orcamentaria": self.uo1.id,
                "unidade_administrativa": self.ua1_uo1.id,
                "unidades_administrativas": [self.ua1_uo1.id, self.ua2_uo1.id],
            }
        )

        response = self.client.post(self.list_url, payload, format="json")

        self.assertEqual(response.status_code, 201)
        criado = User.objects.get(username=payload["username"])
        self.assertEqual(criado.unidade_orcamentaria_id, self.uo1.id)
        self.assertEqual(
            set(criado.unidades_administrativas.values_list("id", flat=True)),
            {self.ua1_uo1.id, self.ua2_uo1.id},
        )
